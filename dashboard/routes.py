"""
Dashboard Blueprint Flask - Endpoints para KPIs e métricas.

Endpoints:
- /api/dashboard/kpis - KPIs agregados
- /api/dashboard/chart/* - Dados para gráficos
- /api/dashboard/alerts/recent - Alertas recentes
- /api/dashboard/sessions/recent-validated - Últimas validadas
"""
from flask import Blueprint, request, jsonify
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

from .service import DashboardService
from backend.database import get_db_context
from sqlalchemy import text

import logging

logger = logging.getLogger(__name__)

# Create blueprint
dashboard_bp = Blueprint('dashboard', __name__, url_prefix='/api/dashboard')


# ============================================================================
# /api/dashboard endpoints
# ============================================================================

@dashboard_bp.route('/kpis', methods=['GET'])
def get_kpis():
    """Retorna KPIs agregados por período."""
    try:
        period = request.args.get('period', 'today')  # today, 7days, 30days, all
        user_id = request.args.get('user_id')  # opcional

        kpis = DashboardService.get_kpis(user_id, period)
        return jsonify({'success': True, 'kpis': kpis})

    except Exception as e:
        logger.error(f"❌ Get KPIs error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/chart/products-per-hour', methods=['GET'])
def get_products_per_hour():
    """Retorna dados para gráfico de produtos por hora."""
    try:
        date_str = request.args.get('date')  # formato: YYYY-MM-DD (opcional, default hoje)

        target_date = None
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        data = DashboardService.get_products_per_hour(target_date)
        return jsonify({'success': True, 'data': data})

    except Exception as e:
        logger.error(f"❌ Get products per hour error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/chart/sessions-per-bay', methods=['GET'])
def get_sessions_per_bay():
    """Retorna dados para gráfico de sessões por baia."""
    try:
        user_id = request.args.get('user_id')
        limit = int(request.args.get('limit', 10))

        data = DashboardService.get_sessions_per_bay(user_id, limit)
        return jsonify({'success': True, 'data': data})

    except Exception as e:
        logger.error(f"❌ Get sessions per bay error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/chart/confidence-distribution', methods=['GET'])
def get_confidence_distribution():
    """Retorna distribuição de confiança das detecções."""
    try:
        data = DashboardService.get_confidence_distribution()
        return jsonify({'success': True, 'data': data})

    except Exception as e:
        logger.error(f"❌ Get confidence distribution error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/alerts/recent', methods=['GET'])
def get_recent_alerts():
    """Retorna alertas recentes."""
    try:
        limit = int(request.args.get('limit', 10))
        alerts = DashboardService.get_recent_alerts(limit)
        return jsonify({'success': True, 'alerts': alerts})

    except Exception as e:
        logger.error(f"❌ Get recent alerts error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@dashboard_bp.route('/sessions/recent-validated', methods=['GET'])
def get_recent_validated_sessions():
    """Retorna últimas sessões validadas."""
    try:
        user_id = request.args.get('user_id')
        limit = int(request.args.get('limit', 5))

        sessions = DashboardService.get_recent_validated_sessions(user_id, limit)
        return jsonify({'success': True, 'sessions': sessions})

    except Exception as e:
        logger.error(f"❌ Get recent validated sessions error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# Excel Export (Complemento ao /api/sessions/export)
# ============================================================================

@dashboard_bp.route('/export/excel', methods=['GET'])
def export_excel():
    """
    Exporta sessões para arquivo Excel com 5 abas formatadas.

    Query params:
    - from_date: Data inicial (YYYY-MM-DD)
    - to_date: Data final (YYYY-MM-DD)

    Retorna: arquivo .xlsx para download
    """
    try:
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        # Build date filter
        date_clause = "1=1"
        if from_date and to_date:
            date_clause = f"started_at >= '{from_date}' AND started_at <= '{to_date}'"
        elif from_date:
            date_clause = f"started_at >= '{from_date}'"
        elif to_date:
            date_clause = f"started_at <= '{to_date}'"

        # Fetch data from database
        with get_db_context() as db:
            # Sessions data
            sessions_query = f"""
                SELECT
                    id, truck_plate, bay_id, camera_id,
                    product_count, ai_count, operator_count,
                    started_at, ended_at, duration_seconds,
                    status, validated_by, validated_at
                FROM counting_sessions
                WHERE {date_clause}
                ORDER BY started_at DESC
            """

            sessions_result = db.execute(text(sessions_query))
            sessions = [row for row in sessions_result.fetchall()]

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ─────────────────────────────────────────────────────────────
        # ABA 1: RESUMO
        # ─────────────────────────────────────────────────────────────
        ws_summary = wb.create_sheet("Resumo")

        # KPIs
        total_sessions = len(sessions)
        total_products = sum(s[4] or 0 for s in sessions)
        total_validated = sum(1 for s in sessions if s[9] == 'validated')
        avg_duration = sum(s[7] or 0 for s in sessions) / total_sessions if total_sessions > 0 else 0

        summary_data = [
            ["Relatório EPI Monitor", ""],
            ["", ""],
            ["Período", f"{from_date or 'Início'} até {to_date or 'Fim'}"],
            ["Total de Sessões", total_sessions],
            ["Total de Produtos", total_products],
            ["Sessões Validadas", total_validated],
            ["Duração Média (min)", f"{avg_duration / 60:.1f}"],
            ["", ""],
            ["Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]

        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx <= 2:
                    cell.font = Font(bold=True, size=14)
                elif row_idx > 2 and col_idx == 1:
                    cell.font = Font(bold=True)

        # ─────────────────────────────────────────────────────────────
        # ABA 2: SESSÕES
        # ─────────────────────────────────────────────────────────────
        ws_sessions = wb.create_sheet("Sessões")

        headers = ["Data", "Baia", "Placa", "Contagem IA", "Contagem Operador", "Diferença", "Duração (min)", "Status", "Validado Por"]
        ws_sessions.append(headers)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws_sessions.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for session in sessions:
            (
                session_id, truck_plate, bay_id, camera_id,
                product_count, ai_count, operator_count,
                started_at, ended_at, duration_seconds,
                status, validated_by, validated_at
            ) = session

            row_data = [
                started_at.strftime("%d/%m/%Y %H:%M") if started_at else "-",
                bay_id or camera_id or "-",
                truck_plate or "-",
                ai_count or 0,
                operator_count or "-",
                (operator_count - ai_count) if operator_count and ai_count else "-",
                f"{(duration_seconds or 0) / 60:.1f}",
                status or "-",
                validated_by or "-"
            ]

            ws_sessions.append(row_data)

            # Apply borders to data rows
            for col_idx in range(1, len(headers) + 1):
                ws_sessions.cell(row=ws_sessions.max_row, column=col_idx).border = border

        # ─────────────────────────────────────────────────────────────
        # ABA 3: ALERTAS (placeholder)
        # ─────────────────────────────────────────────────────────────
        ws_alerts = wb.create_sheet("Alertas")

        ws_alerts.append(["Data/Hora", "Câmera", "Tipo", "Mensagem"])
        ws_alerts.cell(row=1, column=1).font = header_font
        ws_alerts.cell(row=1, column=1).fill = header_fill

        ws_alerts.append(["-", "-", "info", "Sistema de alertas em desenvolvimento"])

        # ─────────────────────────────────────────────────────────────
        # ABA 4: POR HORA
        # ─────────────────────────────────────────────────────────────
        ws_hourly = wb.create_sheet("Por Hora")

        # Aggregate products by hour
        hourly_data = {}
        for session in sessions:
            if session[7]:  # started_at
                hour = session[7].hour
                products = session[4] or 0  # product_count
                hourly_data[hour] = hourly_data.get(hour, 0) + products

        ws_hourly.append(["Hora", "Produtos Contados"])
        ws_hourly.cell(row=1, column=1).font = header_font
        ws_hourly.cell(row=1, column=1).fill = header_fill

        for hour in sorted(hourly_data.keys()):
            ws_hourly.append([f"{hour}:00", hourly_data[hour]])

        # ─────────────────────────────────────────────────────────────
        # ABA 5: POR BAIA
        # ─────────────────────────────────────────────────────────────
        ws_bays = wb.create_sheet("Por Baia")

        # Aggregate by bay
        bay_data = {}
        for session in sessions:
            bay = session[2] or session[3] or "Não identificada"  # bay_id or camera_id
            bay_data[bay] = bay_data.get(bay, 0) + 1

        ws_bays.append(["Baia", "Sessões", "Produtos", "Tempo Médio (min)"])
        ws_bays.cell(row=1, column=1).font = header_font
        ws_bays.cell(row=1, column=1).fill = header_fill

        for bay in sorted(bay_data.keys()):
            bay_sessions = [s for s in sessions if (s[2] or s[3] or "Não identificada") == bay]
            total_products = sum(s[4] or 0 for s in bay_sessions)
            avg_time = sum(s[8] or 0 for s in bay_sessions) / len(bay_sessions) / 60

            ws_bays.append([bay, bay_data[bay], total_products, f"{avg_time:.1f}"])

        # ─────────────────────────────────────────────────────────────
        # GENERATE FILE
        # ─────────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"EPI_Monitor_Relatorio_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        return output.read(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

    except Exception as e:
        logger.error(f"❌ Export Excel error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
